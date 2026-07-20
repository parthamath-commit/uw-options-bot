"""
persistence/excel.py
====================
Signal logger — per-day CSV hot path + on-demand styled xlsx export.

WHY THE REWRITE (2026-07-20):
  The old logger did load_workbook() -> append -> save() for EVERY
  signal (~132/cycle). openpyxl rewrites the entire file on save, so
  each signal cost ~20-25s on the 1-vCPU Oracle Micro VM, stretching
  scan cycles from ~5 min to ~50 min (py-spy confirmed 3/3 stack dumps
  inside openpyxl save/load). A process restart mid-save also corrupted
  the workbook repeatedly (signals_corrupt_*.xlsx).

DESIGN:
  log_signal(sig)  -> appends ONE line to today's signals_YYYY-MM-DD.csv
                      (microseconds, append-only, corruption-proof).
                      Same signature as before — scanner.py unchanged.
  export_xlsx()    -> rebuilds the styled signals.xlsx from TODAY's CSV.
                      Atomic (tmp file + os.replace), so a kill mid-save
                      can never corrupt the visible file.

ROTATION (added 2026-07-20):
  CSV filename carries the date, so each trading day starts a fresh,
  small file. The hot path and the export always touch only today's
  rows — export stays ~1s indefinitely instead of growing unbounded.
  History is preserved as dated CSVs (archive or delete at will);
  cleanup_old_csvs() optionally trims files older than retention_days.

Colour coding (unchanged):
  Green  = bullish (not blocked)
  Red    = bearish (not blocked)
  Amber  = GEX regime blocked
  Navy header
"""

import csv
import logging
import os
from datetime import date, datetime, timedelta
from pathlib import Path

from config import cfg
from models import ScoredSignal

log = logging.getLogger("UWBot.Excel")

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    log.warning("openpyxl not installed — run: pip install openpyxl")


class ExcelLogger:
    COLUMNS = [
        "Timestamp", "Symbol", "Strike", "Right", "Expiry", "Signal Type", "Rule Type", "ML Validation",
        "Structure", "Ask Side", "Intent", "Premium ($)",
        "UW Score", "Additive", "Institutional", "Composite",
        "IV Pct", "Dark Pool", "GEX ($M)", "DEX ($M)",
        "Regime", "Gamma Flip", "Call Wall", "Put Wall",
        "Flow Dir", "GEX Blocked",
        "Bid", "Ask", "Delta",
        "Contracts", "Max Risk $", "Target Premium", "Stop Premium",
    ]
    WIDTHS = [
        22, 8, 9, 6, 12, 14, 12, 14,
        10, 9, 10, 13,
        9, 9, 13, 10,
        8, 12, 10, 10,
        18, 11, 11, 11,
        12, 11,
        7, 7, 7,
        10, 12, 15, 13,
    ]

    _IDX_INTENT  = COLUMNS.index("Intent")
    _IDX_BLOCKED = COLUMNS.index("GEX Blocked")

    # Keep this many days of dated CSVs; older ones removed by cleanup_old_csvs().
    RETENTION_DAYS = 90

    def __init__(self):
        self.path = Path(cfg.EXCEL_PATH)                 # signals.xlsx (view file, always "today")
        self.dir  = self.path.parent
        self.stem = self.path.stem                        # e.g. "signals"
        self.dir.mkdir(parents=True, exist_ok=True)
        self._dirty = False
        self._csv_day = None                              # date the cached csv_path belongs to
        self._csv_path = None
        if OPENPYXL_OK:
            self._hdr  = PatternFill("solid", fgColor="1F3864")
            self._bull = PatternFill("solid", fgColor="C6EFCE")
            self._bear = PatternFill("solid", fgColor="FFC7CE")
            self._blok = PatternFill("solid", fgColor="FFEB9C")
        self._ensure_csv()

    # ── Per-day CSV path ───────────────────────────────────────────────────────
    def _csv_for(self, d: date) -> Path:
        return self.dir / "{}_{}.csv".format(self.stem, d.isoformat())

    @property
    def csv_path(self) -> Path:
        """Today's CSV path. Rolls over automatically at midnight (local time)."""
        today = date.today()
        if self._csv_day != today:
            self._csv_day = today
            self._csv_path = self._csv_for(today)
            self._dirty = False   # new day, fresh export state
        return self._csv_path

    def _ensure_csv(self):
        """Create today's CSV with a header row if missing or empty."""
        try:
            p = self.csv_path
            if not p.exists() or p.stat().st_size == 0:
                with open(p, "w", newline="", encoding="utf-8") as f:
                    csv.writer(f).writerow(self.COLUMNS)
                log.info("Created: {}".format(p))
        except Exception as e:
            log.error("CSV init error: {}".format(e))

    # ── Hot path ───────────────────────────────────────────────────────────────
    def log_signal(self, sig: ScoredSignal, _retry: bool = True):
        """
        Append one signal to today's CSV. Microseconds; never loads the xlsx.
        Signature unchanged (scanner.py and tests call log_signal(sig)).
        """
        row = [
            sig.timestamp,
            sig.symbol,
            sig.strike,
            sig.right,
            sig.expiry,
            getattr(sig, "signal_type",   "unidirectional"),
            getattr(sig, "rule_type",     "unidirectional"),
            getattr(sig, "ml_validation", "rule_only"),
            sig.structure,
            "YES" if sig.ask_side else "NO",
            sig.intent,
            sig.premium,
            sig.uw_score,
            sig.additive_score,
            sig.institutional_score,
            sig.composite_score,
            sig.iv_percentile,
            sig.darkpool_sentiment,
            round(sig.dealer_gex / 1_000_000, 2),
            round(sig.dealer_dex / 1_000_000, 2),
            sig.dealer_regime,
            sig.gamma_flip,
            sig.call_wall,
            sig.put_wall,
            sig.flow_direction,
            "YES" if sig.gex_regime_blocked else "NO",
            sig.live_bid,
            sig.live_ask,
            sig.live_delta,
            sig.suggested_contracts,
            sig.max_loss_dollar,
            sig.target_exit_premium,
            sig.stop_premium,
        ]
        try:
            self._ensure_csv()
            with open(self.csv_path, "a", newline="", encoding="utf-8") as f:
                csv.writer(f).writerow(row)
            self._dirty = True
        except Exception as e:
            log.error("CSV write error: {}".format(e))

    # ── Styled xlsx export (call once per cycle / on demand) ─────────────────
    def export_xlsx(self, force: bool = False):
        """
        Rebuild signals.xlsx from TODAY's CSV with the original styling.
        Atomic write; no-op unless new rows were logged (or force=True).
        Because it reads only today's file, cost stays ~1s regardless of
        how much history has accumulated in prior dated CSVs.
        """
        if not OPENPYXL_OK:
            return
        if not self._dirty and not force:
            return
        try:
            with open(self.csv_path, "r", newline="", encoding="utf-8") as f:
                rows = list(csv.reader(f))
            if not rows:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Signals"

            ws.append(self.COLUMNS)
            for i, (cell, w) in enumerate(zip(ws[1], self.WIDTHS), 1):
                cell.fill      = self._hdr
                cell.font      = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            ws.freeze_panes = "A2"

            for row in rows[1:]:
                ws.append(row)
                blocked = len(row) > self._IDX_BLOCKED and row[self._IDX_BLOCKED] == "YES"
                intent  = row[self._IDX_INTENT] if len(row) > self._IDX_INTENT else ""
                fill = (
                    self._blok if blocked
                    else self._bull if intent == "bullish"
                    else self._bear
                )
                for cell in ws[ws.max_row]:
                    cell.fill = fill

            tmp = self.path.with_name(self.path.stem + ".tmp.xlsx")
            wb.save(tmp)
            os.replace(tmp, self.path)
            self._dirty = False
            log.info("Exported {} signals -> {}".format(len(rows) - 1, self.path))
        except Exception as e:
            log.error("xlsx export error: {}".format(e))

    # ── Retention ──────────────────────────────────────────────────────────────
    def cleanup_old_csvs(self, retention_days: int = None):
        """
        Delete dated CSVs older than retention_days. Safe to call once per
        day (e.g. at market close). Never touches today's file. Silently
        skips files whose names don't parse as a date.
        """
        keep = retention_days if retention_days is not None else self.RETENTION_DAYS
        cutoff = date.today() - timedelta(days=keep)
        prefix = self.stem + "_"
        removed = 0
        try:
            for p in self.dir.glob("{}*.csv".format(prefix)):
                datepart = p.stem[len(prefix):]
                try:
                    d = datetime.strptime(datepart, "%Y-%m-%d").date()
                except ValueError:
                    continue
                if d < cutoff:
                    try:
                        p.unlink()
                        removed += 1
                    except Exception as e:
                        log.warning("Could not remove {}: {}".format(p, e))
            if removed:
                log.info("Cleaned {} CSV(s) older than {} days".format(removed, keep))
        except Exception as e:
            log.error("CSV cleanup error: {}".format(e))